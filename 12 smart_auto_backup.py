import os
import time
import json
import zipfile
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook

# Configuration
MONITOR_FOLDER = r"D:\Z CODE\Z DEVELOPMENT\WEB\Own\0"
BACKUP_FOLDER = r"D:\Z CODE\BACKUP"
LOG_FILE = os.path.join(BACKUP_FOLDER, "backup_log.xlsx")
STATE_FILE = os.path.join(BACKUP_FOLDER, "backup_state.json")
MAIN_BACKUP_ZIP = os.path.join(BACKUP_FOLDER, "main_backup.zip")
EXCLUDE_LIST = ["__pycache__", ".git", "node_modules", "venv", "build"]

# ----------------- Utility: Human-Readable Size -----------------
def human_readable_size(size_bytes):
    """Convert bytes to KB, MB, or GB (readable string)."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024**2:
        return f"{size_bytes / 1024:.2f} KB"
    elif size_bytes < 1024**3:
        return f"{size_bytes / 1024**2:.2f} MB"
    else:
        return f"{size_bytes / 1024**3:.2f} GB"

# ----------------- Logging -----------------
def log_backup(file_name, backup_name, file_size):
    """Write a log entry to Excel."""
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["SNO", "FILE NAME", "BACKUP NAME", "SIZE", "DATE", "TIME"])
    else:
        wb = load_workbook(LOG_FILE)
        ws = wb.active

    sno = ws.max_row
    now = datetime.now()
    time_str = now.strftime("%I:%M:%S %p")  # 12-hour format with AM/PM

    ws.append([
        sno,
        file_name,
        backup_name,
        human_readable_size(file_size),
        now.strftime("%d-%m-%Y"),
        time_str,
    ])
    wb.save(LOG_FILE)

# ----------------- State Load/Save -----------------
def load_backup_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r") as f:
            return json.load(f)
    return {}

def save_backup_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=4)

# ----------------- Main Incremental Backup -----------------
def update_backup_zip():
    """Perform a smart incremental backup into a single ZIP file."""
    print("\n[+] Checking for new or updated files...")

    prev_state = load_backup_state()
    new_state = {}
    files_to_update = []
    changed_top_folders = set()

    for foldername, subfolders, filenames in os.walk(MONITOR_FOLDER):
        if any(ex in foldername for ex in EXCLUDE_LIST):
            continue
        for filename in filenames:
            file_path = os.path.join(foldername, filename)
            if any(ex in file_path for ex in EXCLUDE_LIST):
                continue

            try:
                mtime = os.path.getmtime(file_path)
            except FileNotFoundError:
                continue

            rel_path = os.path.relpath(file_path, MONITOR_FOLDER)
            new_state[rel_path] = mtime

            if rel_path not in prev_state or mtime > prev_state[rel_path]:
                files_to_update.append((file_path, rel_path))
                parts = rel_path.split(os.sep)
                if len(parts) > 1:
                    changed_top_folders.add(parts[0])

    first_backup = not os.path.exists(MAIN_BACKUP_ZIP)

    if not files_to_update and not first_backup:
        print("[✔] No new or modified files found. Backup is up to date.")
        return

    if first_backup:
        print("[+] Performing first full backup (all files)...")
    else:
        print(f"[+] Updating {len(files_to_update)} file(s) in backup...")

    if not os.path.exists(MAIN_BACKUP_ZIP):
        with zipfile.ZipFile(MAIN_BACKUP_ZIP, "w") as zf:
            pass

    if first_backup:
        with zipfile.ZipFile(MAIN_BACKUP_ZIP, "w", zipfile.ZIP_DEFLATED) as zipf:
            for foldername, subfolders, filenames in os.walk(MONITOR_FOLDER):
                if any(ex in foldername for ex in EXCLUDE_LIST):
                    continue
                for filename in filenames:
                    file_path = os.path.join(foldername, filename)
                    if any(ex in file_path for ex in EXCLUDE_LIST):
                        continue
                    rel_path = os.path.relpath(file_path, MONITOR_FOLDER)
                    zipf.write(file_path, rel_path)
        log_name = "All Folders Backup"
    else:
        with zipfile.ZipFile(MAIN_BACKUP_ZIP, "a", zipfile.ZIP_DEFLATED) as zipf:
            for file_path, rel_path in files_to_update:
                zipf.write(file_path, rel_path)
                print(f"   → Updated: {rel_path}")
        log_name = ", ".join(sorted(changed_top_folders)) if changed_top_folders else "Unknown Folder"

    save_backup_state(new_state)
    file_size = os.path.getsize(MAIN_BACKUP_ZIP)
    log_backup(log_name, os.path.basename(MAIN_BACKUP_ZIP), file_size)
    print(f"[✔] Backup {'created' if first_backup else 'updated'} successfully! ({human_readable_size(file_size)})")

# ----------------- Watchdog Handler -----------------
class BackupHandler(FileSystemEventHandler):
    def __init__(self):
        super().__init__()
        self.last_backup_time = 0
        self.cooldown = 10  # seconds to prevent spam

    def on_any_event(self, event):
        if any(ex in event.src_path for ex in EXCLUDE_LIST):
            return
        current_time = time.time()
        if current_time - self.last_backup_time > self.cooldown:
            update_backup_zip()
            self.last_backup_time = current_time

# ----------------- Main -----------------
if __name__ == "__main__":
    print("🚀 Smart Incremental Backup System (Single ZIP + Smart Logging + Readable Size)")
    print(f"Monitoring Folder: {MONITOR_FOLDER}")
    print(f"Backup File: {MAIN_BACKUP_ZIP}")
    print(f"Excluded: {', '.join(EXCLUDE_LIST)}")
    print(f"Logs: {LOG_FILE}")
    print("--------------------------------------------------")

    event_handler = BackupHandler()
    observer = Observer()
    observer.schedule(event_handler, MONITOR_FOLDER, recursive=True)
    observer.start()

    try:
        while True:
            time.sleep(2)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
