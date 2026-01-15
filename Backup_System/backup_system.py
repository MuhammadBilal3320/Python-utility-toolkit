# ---------------------------------------------------------------
# Image Backup and Logger System
# Author: Muhammad Bilal
# Purpose: Create binary backups of files and log them in Excel
# ---------------------------------------------------------------

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


# *******************************************************
#     Function to create Excel file if it doesn't exist
# *******************************************************
def create_excel_if_not_exist(excel_file):
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "BACKUP LOG"
        ws.append(["SNO", "FILE NAME", "BACKUP NAME", "FILE SIZE", "DATE & TIME"])
        wb.save(excel_file)
        print(f"Created new Excel log file: {excel_file}")
        

# *******************************************************
#     Function to Log Backup information into Excel
# *******************************************************
def log_backup(excel_file, file_name, backup_name, file_size):
    wb = load_workbook(excel_file)
    ws = wb.active

    # --- Step 1: Locate the header cell that says "SNO" ---
    header_row = None
    sno_col = None
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=30):
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == "SNO":
                header_row = cell.row
                sno_col = cell.column
                break
        if header_row and sno_col:
            break

    # --- Step 2: If header not found, default to top-left ---
    if not header_row or not sno_col:
        header_row, sno_col = 1, 1

    # --- Step 3: Find next empty row below the header ---
    next_row = header_row + 1
    while ws.cell(row=next_row, column=sno_col).value:
        next_row += 1

    # --- Step 4: Compute SNO ---
    sno = next_row - header_row

    # --- Step 5: Write data into correct columns (aligned with header) ---
    ws.cell(row=next_row, column=sno_col, value=sno)
    ws.cell(row=next_row, column=sno_col + 1, value=file_name)
    ws.cell(row=next_row, column=sno_col + 2, value=backup_name)
    
    # --- Step 4.1: Format file size to include unit (KB, MB, or GB) ---
    if file_size < 1024:
        size_str = f"{file_size} Bytes"
    elif file_size < 1024 * 1024:
        size_str = f"{round(file_size / 1024, 2)} KB"
    elif file_size < 1024 * 1024 * 1024:
        size_str = f"{round(file_size / (1024 * 1024), 2)} MB"
    else:
        size_str = f"{round(file_size / (1024 * 1024 * 1024), 2)} GB"

    ws.cell(row=next_row, column=sno_col + 3, value=size_str)
    ws.cell(row=next_row, column=sno_col + 4, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    wb.save(excel_file)
    print("Log inserted inside the BACKUP LOG table!\n")


# *******************************************************
#     Function to copy binary file safely
# *******************************************************
def backup_file(file_name):
    
    if not os.path.exists(file_name):
        print(f"File not Found! Please Check the file name and try again.")
        return
    
    name, ext = os.path.splitext(file_name)
    backup_name = f"{name}_backup{ext}"
    
    with open(file_name, 'rb') as source_file, open(backup_name, 'wb') as destination:
        while True:
            chunk = source_file.read(1024*1024)
            if not chunk:
                break
            destination.write(chunk)
    
    file_size = os.path.getsize(file_name)
    print(f"Backup Created Successfully!({file_size/1024:.2f})KB")
    
    log_backup("backup_log.xlsx", file_name, backup_name, file_size)

def main():
    print(f"****************************************************")
    print(f"*                  BACKUP SYSTEM                   *")
    print(f"****************************************************")
    
    excel_file = "backup_log.xlsx"
    
    create_excel_if_not_exist(excel_file)
    
    file_name = input("Enter the file Name (with extension, eg. photo.jpg): ").strip()
    
    backup_file(file_name)
    
if __name__ == "__main__":
    main()